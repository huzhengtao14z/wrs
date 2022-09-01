import openpyxl
from scipy import signal
from scipy.signal import butter, lfilter, freqz
import matplotlib.pyplot as plt
import numpy as np
from sklearn.metrics import mean_squared_error

from math import sqrt


# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("unsup3.xlsx")
rows = []

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        rows.append(col[row].value)

# Create/view notch filter
samp_freq = 1000  # Sample frequency (Hz)
notch_freq = 60.0  # Frequency to be removed from signal (Hz)
quality_factor = 20.0  # Quality factor

# Design a notch filter using signal.iirnotch
b_notch, a_notch = signal.iirnotch(notch_freq, quality_factor, samp_freq)
# Compute magnitude response of the designed filter
freq, h = freqz(b_notch, a_notch, fs=2 * np.pi)

fig = plt.figure(figsize=(8, 6))

# Plot magnitude response of the filter
plt.plot(freq * samp_freq / (2 * np.pi), 20 * np.log10(abs(h)),
         'r', label='Bandpass filter', linewidth='2')

plt.xlabel('Frequency [Hz]', fontsize=20)
plt.ylabel('Magnitude [dB]', fontsize=20)
plt.title('Notch Filter', fontsize=20)
plt.grid()

noisySignal = [x for x in rows if abs(x) < 0.65]
abs_noisySignal = [abs(x) for x in rows if abs(x) < 0.65]
t= [0.001*x for x in range(len(noisySignal))]

# x = np.linspace(0, 1, 1000)  # Generate 1000 sample sequence in 1 sec

# Generate the signal containing f1 and f2
# noisySignal = np.sin(2*np.pi*60*x)
# Apply notch filter to the noisy signal using signal.filtfilt
outputSignal = signal.filtfilt(b_notch, a_notch, abs_noisySignal)
def butter_lowpass(cutoff, fs, order=5):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = butter(order, normal_cutoff, btype='low', analog=False)
    return b, a

def butter_lowpass_filter(data, cutoff, fs, order=5):
    b, a = butter_lowpass(cutoff, fs, order=order)
    y = lfilter(b, a, data)
    return y


# Setting standard filter requirements .
order = 2
fs = 1000
cutoff =  12

b, a = butter_lowpass(cutoff, fs, order)

# Plotting the frequency response.
w, h = freqz(b, a, worN=8000)
plt.subplot(2, 1, 1)
plt.plot(0.5*fs*w/np.pi, np.abs(h), 'b')
plt.plot(cutoff, 0.5*np.sqrt(2), 'ko')
plt.axvline(cutoff, color='k')
#plt.xlim(0, 0.5*fs)
plt.title("Lowpass Filter Frequency Response")
plt.xlabel('Frequency [Hz]')
plt.grid()



# Filtering and plotting
outputSignal = butter_lowpass_filter(outputSignal, cutoff, fs, order)
window = 50
# for i in range()
# rms = sqrt(mean_squared_error(valid_y, pred))

# ave_outputSignal = []
# for i in range(len(outputSignal)-window):
#     a = 0
#     for i in range(i,i+50):
#         a += outputSignal[i]*outputSignal[i]
#     a = sqrt(a/window)
#     ave_outputSignal.append(a)
# np.asarray(ave_outputSignal)

# Plotting
fig = plt.figure(figsize=(8, 6))

plt.subplot(211)
plt.ylim(-0.13,0.13)
plt.plot(t[:-50], noisySignal[:-50], color='b', linewidth=1)
plt.xlabel('Time', fontsize=14)
plt.ylabel('Magnitude', fontsize=14)
plt.title('Raw Signal', fontsize=14)

# Plot notch-filtered version of signal
plt.subplot(212)
plt.ylim(0,0.13)
# Plot output signal of notch filter
plt.plot(t[:-50], abs_noisySignal[:-50], color='b', linewidth=1, alpha=0.2)
plt.plot(t[:-50], outputSignal[:-50], color='r', linewidth=2)
plt.xlabel('Time', fontsize=14)
plt.ylabel('Magnitude', fontsize=14)
plt.title('Rectified and Filtered Signal', fontsize=14)
plt.subplots_adjust(hspace=0.5)
fig.tight_layout()

# Average
# plt.subplot(313)
# ave_outputSignal = []
# for i in range(len(outputSignal)-30):
#     ave_outputSignal.append(np.average(outputSignal[i:i+30]))
# np.asarray(ave_outputSignal)
# plt.plot(ave_outputSignal)
# plt.xlabel('Time', fontsize=14)
# plt.ylabel('Magnitude', fontsize=14)
# plt.title('Filtered Signal', fontsize=14)
# plt.ylim(-0.15,0.15)
plt.show()

from scipy.fft import fft, fftfreq
# Number of sample points
N = len(outputSignal)
# sample spacing
T = 1 / 1000
y = noisySignal
yf = fft(y)
xf = fftfreq(N, T)[:N//2]
import matplotlib.pyplot as plt
plt.plot(xf, 2.0/N * np.abs(yf[0:N//2]))
plt.grid()
plt.show()

